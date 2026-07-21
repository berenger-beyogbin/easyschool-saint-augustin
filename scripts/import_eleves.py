"""Import sécurisé de l'ancien export Excel ELEVE dans la table Eleve.

Sans --commit, l'import est simulé puis annulé. Les anciens IDEleve sont
conservés par défaut afin de permettre l'import ultérieur des inscriptions.
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


EXPECTED_HEADERS = [
    "IDEleve", "Matricule", "nom", "Prenoms", "datnais", "lieunais",
    "sexe", "Nationalite", "Religion", "IDFamille", "photo", "numExtr",
    "datExtr", "lieuExtr", "Habitation", "NomUrg", "CntactUrg", "HabitUrg",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def as_int(value: Any) -> int | None:
    raw = clean(value)
    if not raw:
        return None
    try:
        return int(float(raw.replace(",", ".")))
    except ValueError:
        return None


def as_optional_id(value: Any) -> int | None:
    result = as_int(value)
    return result if result is not None and result > 0 else None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            result = converted.date() if isinstance(converted, datetime) else converted
            return result if isinstance(result, date) and 1900 <= result.year <= date.today().year else None
        except (OverflowError, ValueError):
            return None
    raw = clean(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            result = datetime.strptime(raw, fmt).date()
            return result if 1900 <= result.year <= date.today().year else None
        except ValueError:
            pass
    return None


def repaired_workbook(source: Path, temp_dir: Path) -> Path:
    """Corrige dans une copie temporaire la faute XML `biltinId` de l'export."""
    repaired = temp_dir / "eleves_repare.xlsx"
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(repaired, "w") as outgoing:
        for item in incoming.infolist():
            data = incoming.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = data.replace(b"biltinId", b"builtinId")
            outgoing.writestr(item, data)
    return repaired


def read_rows(source: Path) -> list[tuple[int, dict[str, Any]]]:
    with tempfile.TemporaryDirectory(prefix="easy_school_eleves_") as folder:
        repaired = repaired_workbook(source, Path(folder))
        workbook = load_workbook(repaired, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(iterator)]
        missing = [name for name in EXPECTED_HEADERS if name not in headers]
        if missing:
            raise ValueError(f"Colonnes absentes : {', '.join(missing)}")
        rows = [(number, dict(zip(headers, values))) for number, values in enumerate(iterator, start=2)]
        workbook.close()
        return rows


def payload(source: dict[str, Any], preserve_ids: bool) -> dict[str, Any]:
    data = {
        "Matricule": clean(source.get("Matricule"))[:50],
        "Nom": clean(source.get("nom"))[:35],
        "Prenoms": clean(source.get("Prenoms"))[:75],
        "DateNaissance": as_date(source.get("datnais")),
        "LieuNaissance": clean(source.get("lieunais"))[:30] or None,
        "Sexe": as_int(source.get("sexe")),
        "IDNationalite": as_optional_id(source.get("Nationalite")),
        "IDReligion": as_optional_id(source.get("Religion")),
        "IDFamille": as_optional_id(source.get("IDFamille")),
        "PhotoPath": clean(source.get("photo"))[:255] or None,
        "NumExtrait": clean(source.get("numExtr"))[:20] or None,
        "DateExtrait": as_date(source.get("datExtr")),
        "LieuDelivrance": clean(source.get("lieuExtr"))[:30] or None,
        "Habitation": clean(source.get("Habitation"))[:25] or None,
        "NomUrgence": clean(source.get("NomUrg"))[:50] or None,
        "ContactUrgence": clean(source.get("CntactUrg"))[:50] or None,
        "HabitationUrgence": clean(source.get("HabitUrg"))[:50] or None,
    }
    if preserve_ids:
        data["IDEleve"] = as_int(source.get("IDEleve"))
    return data


def write_report(path: Path, rejected: list[tuple[int, str, str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["Ligne Excel", "Ancien IDEleve", "Matricule", "Motif"])
        writer.writerows(rejected)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importer Export eleve EPC.xlsx")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--new-ids", action="store_true")
    parser.add_argument("--rejects", type=Path, default=Path("eleves_rejetes.csv"))
    parser.add_argument("--warnings", type=Path, default=Path("eleves_avertissements.csv"))
    args = parser.parse_args()
    if not args.xlsx.is_file():
        parser.error(f"Fichier introuvable : {args.xlsx}")

    from sqlalchemy import select, text
    from sqlalchemy.exc import IntegrityError
    from app.database import get_session
    from models.eleve import Eleve
    from models.famille import TFamille
    from models.nationalite import TNationalite
    from models.religion import TReligion

    preserve_ids = not args.new_ids
    sources = read_rows(args.xlsx)
    candidates: list[tuple[int, dict[str, Any]]] = []
    rejected: list[tuple[int, str, str, str]] = []
    warnings: list[tuple[int, str, str, str]] = []
    seen_matricules: set[str] = set()
    seen_ids: set[int] = set()

    for excel_row, source in sources:
        data = payload(source, preserve_ids)
        reason = ""
        if not data["Matricule"]:
            reason = "Matricule obligatoire manquant"
        elif not data["Nom"]:
            reason = "Nom obligatoire manquant"
        elif not data["Prenoms"]:
            reason = "Prénoms obligatoires manquants"
        elif not data["DateNaissance"]:
            reason = "Date de naissance absente ou invalide"
        elif data["Sexe"] not in (1, 2):
            reason = "Sexe invalide (1 ou 2 attendu)"
        elif data["Matricule"].casefold() in seen_matricules:
            reason = "Matricule dupliqué dans le fichier"
        elif preserve_ids and not data.get("IDEleve"):
            reason = "Ancien IDEleve invalide"
        elif preserve_ids and data["IDEleve"] in seen_ids:
            reason = "Ancien IDEleve dupliqué dans le fichier"
        if reason:
            rejected.append((excel_row, clean(source.get("IDEleve")), data["Matricule"], reason))
            continue
        seen_matricules.add(data["Matricule"].casefold())
        if preserve_ids:
            seen_ids.add(data["IDEleve"])
        candidates.append((excel_row, data))

    session = get_session()
    inserted = 0
    try:
        family_ids = set(session.scalars(select(TFamille.IdTFamille)))
        nationality_ids = set(session.scalars(select(TNationalite.IDTNationalite)))
        religion_ids = set(session.scalars(select(TReligion.IDTReligion)))
        existing_matricules = {value.casefold() for value in session.scalars(select(Eleve.Matricule))}
        existing_ids = set(session.scalars(select(Eleve.IDEleve))) if preserve_ids else set()

        for excel_row, data in candidates:
            reason = ""
            if data["Matricule"].casefold() in existing_matricules:
                reason = "Matricule déjà présent dans la base"
            elif preserve_ids and data["IDEleve"] in existing_ids:
                reason = "IDEleve déjà présent dans la base"
            if not reason and data["IDFamille"] is not None and data["IDFamille"] not in family_ids:
                missing = data["IDFamille"]
                data["IDFamille"] = None
                warnings.append((excel_row, clean(data.get("IDEleve")), data["Matricule"], f"Famille {missing} absente : élève importé sans famille"))
            if not reason and data["IDNationalite"] is not None and data["IDNationalite"] not in nationality_ids:
                missing = data["IDNationalite"]
                data["IDNationalite"] = None
                warnings.append((excel_row, clean(data.get("IDEleve")), data["Matricule"], f"Nationalité {missing} absente : valeur laissée vide"))
            if not reason and data["IDReligion"] is not None and data["IDReligion"] not in religion_ids:
                missing = data["IDReligion"]
                data["IDReligion"] = None
                warnings.append((excel_row, clean(data.get("IDEleve")), data["Matricule"], f"Religion {missing} absente : valeur laissée vide"))
            if reason:
                rejected.append((excel_row, clean(data.get("IDEleve")), data["Matricule"], reason))
                continue
            session.add(Eleve(**data))
            existing_matricules.add(data["Matricule"].casefold())
            if preserve_ids:
                existing_ids.add(data["IDEleve"])
            inserted += 1

        session.flush()
        if preserve_ids and inserted:
            session.execute(text(
                "SELECT setval(pg_get_serial_sequence('\"Eleve\"', 'IDEleve'), "
                "COALESCE((SELECT MAX(\"IDEleve\") FROM \"Eleve\"), 1), true)"
            ))
        write_report(args.rejects, rejected)
        write_report(args.warnings, warnings)
        if args.commit:
            session.commit()
            print("IMPORT VALIDÉ")
        else:
            session.rollback()
            print("SIMULATION : aucune modification enregistrée")
        print(f"Lignes lues : {len(sources)}")
        print(f"Lignes prêtes à être insérées : {inserted}")
        print(f"Lignes rejetées/ignorées : {len(rejected)}")
        print(f"Avertissements (ligne importée) : {len(warnings)}")
        print(f"Rapport : {args.rejects.resolve()}")
        print(f"Avertissements : {args.warnings.resolve()}")
        return 0
    except IntegrityError as exc:
        session.rollback()
        print(f"Import annulé : {exc.orig}", file=sys.stderr)
        return 2
    except Exception as exc:
        session.rollback()
        print(f"Import annulé : {exc}", file=sys.stderr)
        return 2
    finally:
        session.close()


if __name__ == "__main__":
    raise SystemExit(main())
