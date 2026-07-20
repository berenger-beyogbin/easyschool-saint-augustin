"""Import sécurisé d'un ancien export Excel FAMILLE dans TFamille.

Exemples :
  python scripts/import_familles.py "C:\\Users\\BBY\\Documents\\Export FAMILLE EPC.xlsx" --analyze-only
  python scripts/import_familles.py "C:\\Users\\BBY\\Documents\\Export FAMILLE EPC.xlsx"
  python scripts/import_familles.py "C:\\Users\\BBY\\Documents\\Export FAMILLE EPC.xlsx" --commit

Sans --commit, la transaction est toujours annulée après simulation.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Permet d'exécuter ce fichier directement depuis le dossier scripts/.
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

EXPECTED_HEADERS = [
    "IdTFamille", "pa_nomResp", "pa_Qual", "pa_Prof", "pa_type",
    "pa_NumPid", "pa_Adr", "pa_Cel", "pa_Mail", "SIMaitre",
    "EbrieAbobote", "NomUrgence", "ContactUrgence", "HabitationUrgence",
    "UrgenceMoimeme", "HabitationParent", "NomPE", "ProfessionPE",
    "TelPE", "NomME", "ProfEssionME", "TélME", "EnsCatPri", "EnsCatSec",
]


@dataclass
class Rejection:
    excel_row: int
    legacy_id: str
    reason: str
    name: str
    phone: str


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def as_int(value: Any, default: int | None = None) -> int | None:
    raw = clean(value)
    if not raw:
        return default
    try:
        return int(float(raw.replace(",", ".")))
    except ValueError:
        return default


def as_bool(value: Any) -> bool:
    return clean(value).lower() in {"1", "true", "vrai", "oui", "yes", "x"}


def normalize_email(value: Any) -> str | None:
    email = clean(value)
    if not email:
        return None
    return email if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email) else None


def choose_name(row: dict[str, Any]) -> str:
    responsible = clean(row.get("pa_nomResp"))
    if responsible:
        return responsible
    quality = as_int(row.get("pa_Qual"), 1)
    father, mother = clean(row.get("NomPE")), clean(row.get("NomME"))
    if quality == 2:
        return father or mother
    if quality == 3:
        return mother or father
    return father or mother or clean(row.get("NomUrgence"))


def choose_phone(row: dict[str, Any]) -> str:
    responsible = clean(row.get("pa_Cel"))
    if responsible:
        return responsible
    quality = as_int(row.get("pa_Qual"), 1)
    father, mother = clean(row.get("TelPE")), clean(row.get("TélME"))
    if quality == 2:
        return father or mother or clean(row.get("ContactUrgence"))
    if quality == 3:
        return mother or father or clean(row.get("ContactUrgence"))
    return father or mother or clean(row.get("ContactUrgence"))


def read_rows(path: Path) -> tuple[list[tuple[int, dict[str, Any]]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    missing = [header for header in EXPECTED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Colonnes absentes de l'export : {', '.join(missing)}")
    rows = [(number, dict(zip(headers, values))) for number, values in enumerate(iterator, start=2)]
    workbook.close()
    return rows, headers


def build_payload(row: dict[str, Any], preserve_ids: bool) -> dict[str, Any]:
    payload = {
        "NomResponsable": choose_name(row)[:50],
        "QualiteResponsable": as_int(row.get("pa_Qual"), 1) or 1,
        "ProfessionResponsable": clean(row.get("pa_Prof"))[:25] or None,
        "TypeResponsable": as_int(row.get("pa_type")),
        "NumeroPieceIdentite": clean(row.get("pa_NumPid"))[:25] or None,
        "AdresseResponsable": clean(row.get("pa_Adr"))[:25] or None,
        "CellulaireResponsable": choose_phone(row)[:25],
        "EmailResponsable": normalize_email(row.get("pa_Mail")),
        "SIMaitre": as_bool(row.get("SIMaitre")),
        "EbrieAbobote": as_bool(row.get("EbrieAbobote")),
        "NomUrgence": clean(row.get("NomUrgence"))[:50] or None,
        "ContactUrgence": clean(row.get("ContactUrgence"))[:20] or None,
        "HabitationUrgence": clean(row.get("HabitationUrgence"))[:50] or None,
        "UrgenceMoimeme": as_bool(row.get("UrgenceMoimeme")),
        "HabitationParent": clean(row.get("HabitationParent"))[:50] or None,
        "NomPere": clean(row.get("NomPE"))[:50] or None,
        "ProfessionPere": clean(row.get("ProfessionPE"))[:25] or None,
        "TelPere": clean(row.get("TelPE"))[:25] or None,
        "NomMere": clean(row.get("NomME"))[:50] or None,
        "ProfessionMere": clean(row.get("ProfEssionME"))[:25] or None,
        "TelMere": clean(row.get("TélME"))[:25] or None,
        "EnsCatPrimaire": as_bool(row.get("EnsCatPri")),
        "EnsCatSecondaire": as_bool(row.get("EnsCatSec")),
    }
    if preserve_ids:
        payload["IdTFamille"] = as_int(row.get("IdTFamille"))
    return payload


def write_rejections(path: Path, rows: list[Rejection]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["Ligne Excel", "Ancien ID", "Motif", "Nom retenu", "Téléphone retenu"])
        writer.writerows((r.excel_row, r.legacy_id, r.reason, r.name, r.phone) for r in rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importer l'ancien export FAMILLE vers TFamille.")
    parser.add_argument("xlsx", type=Path, help="Chemin du fichier Export FAMILLE EPC.xlsx")
    parser.add_argument("--commit", action="store_true", help="Valider réellement l'import en base.")
    parser.add_argument("--analyze-only", action="store_true", help="Analyser le fichier sans connexion à la base.")
    parser.add_argument("--new-ids", action="store_true", help="Générer de nouveaux ID (déconseillé avant import des élèves).")
    parser.add_argument("--rejects", type=Path, default=Path("familles_rejetees.csv"))
    args = parser.parse_args()

    if not args.xlsx.is_file():
        parser.error(f"Fichier introuvable : {args.xlsx}")

    source_rows, _ = read_rows(args.xlsx)
    preserve_ids = not args.new_ids
    accepted: list[tuple[int, dict[str, Any]]] = []
    rejected: list[Rejection] = []
    seen_keys: set[tuple[str, str]] = set()
    seen_ids: set[int] = set()

    for excel_row, source in source_rows:
        payload = build_payload(source, preserve_ids)
        name = payload["NomResponsable"]
        phone = payload["CellulaireResponsable"]
        legacy_id = clean(source.get("IdTFamille"))
        reason = ""
        if not name:
            reason = "Nom du responsable introuvable"
        elif not phone:
            reason = "Téléphone du responsable introuvable"
        elif preserve_ids and not payload.get("IdTFamille"):
            reason = "Ancien IdTFamille invalide"
        key = (name.casefold(), phone.casefold())
        if not reason and key in seen_keys:
            reason = "Doublon NomResponsable + CellulaireResponsable dans le fichier"
        if not reason and preserve_ids and payload["IdTFamille"] in seen_ids:
            reason = "Ancien IdTFamille dupliqué dans le fichier"
        if reason:
            rejected.append(Rejection(excel_row, legacy_id, reason, name, phone))
            continue
        seen_keys.add(key)
        if preserve_ids:
            seen_ids.add(payload["IdTFamille"])
        accepted.append((excel_row, payload))

    if args.analyze_only:
        write_rejections(args.rejects, rejected)
        print(f"Lignes lues : {len(source_rows)}")
        print(f"Importables : {len(accepted)}")
        print(f"Rejetées : {len(rejected)}")
        print(f"Rapport : {args.rejects.resolve()}")
        return 0

    from sqlalchemy import select, text
    from sqlalchemy.exc import IntegrityError

    from app.database import get_session
    from models.famille import TFamille

    session = get_session()
    inserted = 0
    try:
        existing_keys = {
            (clean(name).casefold(), clean(phone).casefold())
            for name, phone in session.execute(
                select(TFamille.NomResponsable, TFamille.CellulaireResponsable)
            )
        }
        existing_ids = set(session.scalars(select(TFamille.IdTFamille))) if preserve_ids else set()
        for excel_row, payload in accepted:
            key = (payload["NomResponsable"].casefold(), payload["CellulaireResponsable"].casefold())
            legacy_id = clean(payload.get("IdTFamille"))
            if key in existing_keys:
                rejected.append(Rejection(
                    excel_row, legacy_id, "Déjà présent dans la base",
                    payload["NomResponsable"], payload["CellulaireResponsable"],
                ))
                continue
            if preserve_ids and payload["IdTFamille"] in existing_ids:
                rejected.append(Rejection(
                    excel_row, legacy_id, "ID déjà présent dans la base",
                    payload["NomResponsable"], payload["CellulaireResponsable"],
                ))
                continue
            session.add(TFamille(**payload))
            existing_keys.add(key)
            if preserve_ids:
                existing_ids.add(payload["IdTFamille"])
            inserted += 1

        session.flush()
        if preserve_ids and inserted:
            session.execute(text(
                "SELECT setval(pg_get_serial_sequence('\"TFamille\"', 'IdTFamille'), "
                "COALESCE((SELECT MAX(\"IdTFamille\") FROM \"TFamille\"), 1), true)"
            ))
        write_rejections(args.rejects, rejected)
        if args.commit:
            session.commit()
            action = "IMPORT VALIDÉ"
        else:
            session.rollback()
            action = "SIMULATION : aucune modification enregistrée"
        print(action)
        print(f"Lignes prêtes à être insérées : {inserted}")
        print(f"Lignes rejetées/ignorées : {len(rejected)}")
        print(f"Rapport : {args.rejects.resolve()}")
        return 0
    except IntegrityError as exc:
        session.rollback()
        print(f"Import annulé à cause d'une contrainte de base : {exc.orig}", file=sys.stderr)
        return 2
    except Exception as exc:
        session.rollback()
        print(f"Import annulé : {exc}", file=sys.stderr)
        return 2
    finally:
        session.close()


if __name__ == "__main__":
    raise SystemExit(main())
